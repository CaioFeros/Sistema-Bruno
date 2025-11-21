"""
Módulo para exportação de dados para Excel.
"""
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def format_excel_file(file_path: str, has_stats: bool = False):
    """
    Formata o arquivo Excel gerado com estilos e larguras de coluna.
    
    Args:
        file_path: Caminho para o arquivo Excel
        has_stats: Se True, formata também a aba de estatísticas
    """
    try:
        wb = load_workbook(file_path)
        
        # Formatar aba de Recibos
        if 'Recibos' in wb.sheetnames:
            ws = wb['Recibos']
            _format_receipts_sheet(ws)
        
        # Formatar aba de Estatísticas
        if has_stats and 'Estatísticas por Vendedor' in wb.sheetnames:
            ws_stats = wb['Estatísticas por Vendedor']
            _format_stats_sheet(ws_stats)
        
        wb.save(file_path)
    except Exception as e:
        # Se houver erro na formatação, o arquivo ainda será salvo sem formatação
        print(f"Aviso: Não foi possível formatar o Excel: {str(e)}")


def _format_receipts_sheet(ws):
    """Formata a aba de Recibos com estilo profissional."""
    # Cores para formatação
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    row_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Formatar cabeçalho
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar larguras das colunas
    column_widths = {
        'A': 15,  # Nº Recibo
        'B': 30,  # Vendedor
        'C': 35,  # Cliente
        'D': 40,  # Descrição do Produto
        'E': 12,  # Quantidade
        'F': 15   # Valor Unitário
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Formatar células de dados com linhas intercaladas (zebrado)
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_font = Font(size=10, color="000000")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Alternar cores: linhas pares = branco, linhas ímpares = cinza claro
        fill_color = row_white if row_idx % 2 == 0 else row_gray
        
        for cell in row:
            cell.alignment = data_alignment
            cell.fill = fill_color
            cell.font = data_font
    
    # Formatar coluna de valores (F)
    for cell in ws['F'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Formatar coluna de quantidade (E)
    for cell in ws['E'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '0'
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'


def _format_stats_sheet(ws):
    """Formata a aba de Estatísticas por Vendedor com blocos separados."""
    # Cores para formatação
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    block_fill_light = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    block_fill_dark = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    separator_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Formatar cabeçalho
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar larguras das colunas
    column_widths = {
        'A': 30,  # Vendedor
        'B': 40,  # Produto
        'C': 18,  # Quantidade Total
        'D': 18,  # Valor Total
        'E': 20,  # Preço Médio por MG
        'F': 20,  # Preço Mínimo por MG
        'G': 20   # Preço Máximo por MG
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Identificar blocos por vendedor e formatar
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_font = Font(size=10, color="000000")
    vendedor_font = Font(size=10, color="000000", bold=True)
    
    # Borda para separação entre blocos
    thick_border = Border(
        top=Side(style='medium', color='808080'),
        bottom=Side(style='thin', color='808080')
    )
    
    current_vendedor = None
    block_counter = 0  # Para alternar cores entre blocos
    previous_row_idx = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        vendedor_cell = row[0]  # Primeira coluna é o vendedor
        vendedor_value = vendedor_cell.value if vendedor_cell.value else ""
        
        # Verificar se mudou o vendedor
        if vendedor_value != current_vendedor:
            # Novo bloco de vendedor
            if current_vendedor is not None and previous_row_idx is not None:
                # Adicionar borda superior na primeira linha do novo bloco para separação visual
                for cell in row:
                    cell.border = Border(top=Side(style='medium', color='808080'))
            
            current_vendedor = vendedor_value
            block_counter += 1
        
        # Alternar cores entre blocos: blocos ímpares = claro, blocos pares = mais escuro
        block_fill = block_fill_light if block_counter % 2 == 1 else block_fill_dark
        
        for cell_idx, cell in enumerate(row):
            cell.alignment = data_alignment
            cell.fill = block_fill
            
            # Destacar coluna de vendedor com fonte bold
            if cell_idx == 0:
                cell.font = vendedor_font
            else:
                cell.font = data_font
        
        previous_row_idx = row_idx
    
    # Formatar coluna de Valor Total (D)
    for cell in ws['D'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Formatar coluna de Quantidade Total (C)
    for cell in ws['C'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Formatar coluna de Preço Médio por MG (E)
    for cell in ws['E'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Formatar coluna de Preço Mínimo por MG (F)
    for cell in ws['F'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Formatar coluna de Preço Máximo por MG (G)
    for cell in ws['G'][1:]:  # Pular cabeçalho
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'


def export_to_excel(df: pd.DataFrame, output_dir: Optional[str] = None) -> str:
    """
    Exporta DataFrame para arquivo Excel formatado.
    
    Args:
        df: DataFrame pandas a ser exportado
        output_dir: Diretório onde salvar o arquivo (None = diretório atual)
        
    Returns:
        Caminho do arquivo Excel criado
        
    Raises:
        Exception: Se houver erro ao exportar
    """
    try:
        # Criar nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"recibos_extraidos_{timestamp}.xlsx"
        
        if output_dir:
            output_path = Path(output_dir) / filename
        else:
            output_path = Path(filename)
        
        # Converter valores numéricos de volta para formato brasileiro antes de exportar
        df_export = df.copy()
        
        if 'Valor Unitário' in df_export.columns:
            # Formatar valores como string no formato brasileiro
            def format_brazilian_currency(value):
                if pd.isna(value) or value == 0:
                    return ''
                return f"{value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            df_export['Valor Unitário'] = df_export['Valor Unitário'].apply(format_brazilian_currency)
        
        # Exportar para Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Recibos')
        
        # Formatar o arquivo
        format_excel_file(str(output_path))
        
        return str(output_path)
    
    except Exception as e:
        raise Exception(f"Erro ao exportar para Excel: {str(e)}")


def export_to_excel_with_path(df: pd.DataFrame, file_path: str, df_stats: Optional[pd.DataFrame] = None) -> str:
    """
    Exporta DataFrame para um caminho específico de arquivo Excel.
    Pode incluir uma segunda aba com estatísticas por vendedor.
    
    Args:
        df: DataFrame pandas a ser exportado (aba Recibos)
        file_path: Caminho completo do arquivo Excel a ser criado
        df_stats: DataFrame opcional com estatísticas por vendedor (aba Estatísticas)
        
    Returns:
        Caminho do arquivo Excel criado
        
    Raises:
        Exception: Se houver erro ao exportar
    """
    try:
        # Converter valores numéricos de volta para formato brasileiro antes de exportar
        df_export = df.copy()
        
        if 'Valor Unitário' in df_export.columns:
            # Formatar valores como string no formato brasileiro
            def format_brazilian_currency(value):
                if pd.isna(value) or value == 0:
                    return ''
                return f"{value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            df_export['Valor Unitário'] = df_export['Valor Unitário'].apply(format_brazilian_currency)
        
        # Preparar DataFrame de estatísticas se fornecido
        # Manter valores numéricos para formatação no Excel (não converter para string)
        df_stats_export = df_stats.copy() if (df_stats is not None and not df_stats.empty) else None
        
        # Garantir que o diretório existe
        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Exportar para Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Recibos')
            
            if df_stats_export is not None and not df_stats_export.empty:
                df_stats_export.to_excel(writer, index=False, sheet_name='Estatísticas por Vendedor')
        
        # Formatar o arquivo
        format_excel_file(str(output_path), has_stats=(df_stats_export is not None and not df_stats_export.empty))
        
        return str(output_path)
    
    except Exception as e:
        raise Exception(f"Erro ao exportar para Excel: {str(e)}")

